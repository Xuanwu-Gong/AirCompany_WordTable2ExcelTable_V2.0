# -*- coding: utf-8 -*-
"""
@author: GongWuxuan
"""
import os,sys
import re
import openpyxl
import tkinter
import pandas as pd
from docx import Document
from tkinter import filedialog
 
def write_excel_xlsx(path, sheet_name, value):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.append(value)
    workbook.save(path)

 
def read_excel_xlsx_rows(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    return sheet.rows

#遍历目标目录下的word文件
def print_list_dir(docFiles,Path):
    try:
        filelist = os.listdir(Path)
    except FileNotFoundError:
        print('系统找不到指定的路径')
        sys.exit(0)
    except NotADirectoryError:
        print('目录名称无效')
        sys.exit(0)
        
    for file in  filelist:
        file_path=os.path.join(Path,file)  #路径拼接成绝对路径
        if os.path.isfile(file_path):  #如果是文件，就保存这个文件路径
            if file_path.endswith(".docx"):
                docFiles.append(file_path)
        if os.path.isdir(file_path):  #如果目录，就递归子目录
            print_list_dir(docFiles,file_path)

def next_dir(path):
    list = []
    if (os.path.exists(path)):
        files = os.listdir(path)
        for file in files:
            m = os.path.join(path,file)
            if (os.path.isdir(m)):
                h = os.path.split(m)
                list.append(h[1])
        return list
    
def selectPath():
    Path_ = filedialog.askdirectory()
    Path.set(Path_)

def chinese2digits(chinese_num):#百以内的自然数（不包括一百）转换成阿拉伯数字
    try:
        chinese_num_map ={'零':0, '一':1, '二':2, '两':2, '三':3, '四':4, '五':5, '六':6, '七':7, '八':8, '九':9, '十':10}
        if len(chinese_num) == 1:
            digit = chinese_num_map[chinese_num]
            return digit
        
        elif len(chinese_num) == 2:
            if chinese_num[1] == '十':
                digit = chinese_num_map[chinese_num[0]] * 10
            else:
                digit = chinese_num_map[chinese_num[1]] + 10
            return digit
        
        elif len(chinese_num) == 3:
            digit = chinese_num_map[chinese_num[0]] * 10 + chinese_num_map[chinese_num[2]]
            return digit
        
        else :
            return -1
    
    except KeyError:
        return -1
    
root = tkinter.Tk()
Path = tkinter.StringVar()
tkinter.Label(root,text = "目标路径:").grid(row = 0, column = 0)
tkinter.Entry(root, textvariable = Path).grid(row = 0, column = 1)
tkinter.Button(root, text = "路径选择", command = selectPath).grid(row = 0, column = 2)
root.mainloop()

#Path = input('请输入文件夹路径：')
#Path = "D:\GongWuxuan\Code\Python\002_WordConvExcel"
Path = Path.get()
Path = os.path.abspath(Path)
GenPath = Path + '\\GenExcel'

seq = ['姓名', '日期', '班组', '见习扇区', '教员', '时长', '执照类别', '管制级别', '单位及职务', '科目自我评定', '教员评定'] 
#try:
#    filelist = os.listdir(Path)
#except FileNotFoundError:
#    print('系统找不到指定的路径')
#    sys.exit(0)
#except NotADirectoryError:
#    print('目录名称无效')
#    sys.exit(0)
groups = next_dir(Path)

if not os.path.exists(GenPath):
      os.makedirs(GenPath)
try:
    groups.remove('GenExcel') 
except ValueError:
    groups = groups

f = open(GenPath + '\\' + 'log.txt', 'a')
sys.stdout = f
sys.stderr = f # redirect std err, if necessary

for group in groups:
    group_path = os.path.join(Path,group)
    target_file = GenPath + '\\' + group + '.xlsx'
    crew_names = next_dir(group_path)
    
    null_df = pd.DataFrame(columns=seq,index=[])
    data_writer = pd.ExcelWriter(target_file)
    for crew_name in crew_names:
        null_df.to_excel(data_writer,sheet_name=crew_name, index=False, header=True)
    data_writer.save()
    data_writer.close()
    
    for crew_name in crew_names:
        docFiles = [] #初始化文档列表
        Num = 0 #初始化计数器
        crew_Path = os.path.join(group_path,crew_name)
        print_list_dir(docFiles, crew_Path)

        for docFile in docFiles:
            diction = [] #初始化字典
            information = [] #初始化信息列表
            flag = 0 #初始化flag
            
            try:
                document = Document(docFile)
            except :
                print('无法打开文件' + docFile)
                continue
            
            table = document.tables
            table = table[0] #根据甲方规定，word格式给定，只有一种表格
            
            #读取word文档里面的表格信息
            for i in range(len(table.rows)):
                for j in range(len(table.columns)):
                    information.append(table.cell(i,j).text)
            
            #删除重复信息
            setinfor = sorted(set(information),key= information.index)
            
            #删除空格、查找关键字
            for i in range(len(setinfor)):
                setinfor[i] = setinfor[i].strip(' ')
                setinfor[i] = setinfor[i].strip('\n')
                if flag == 0:
                    find = re.search('复训科目及时间',setinfor[i])
                    if find :
                        try:
                            time = re.search(r'\d\d\d\d*年*\d*月*\d*日',setinfor[i]).group(0)
                            flag = 1
                            find = 0
                        except AttributeError:
                            print('文件' + docFile + ' 找不到复训时间')
                            
                elif flag == 1:
                    find = re.search('科目自我评定',setinfor[i])
                    if find :
                        setinfor[i] = setinfor[i].strip('科目自我评定')
                        setinfor[i] = setinfor[i].strip('\n')
                        setinfor[i] = setinfor[i].strip(' ')
                        selfCommentNum = i
                        flag = 2
                        find = 0
                        
                elif flag == 2:
                    find = re.search('教员评定',setinfor[i])
                    if find :
                        try:
                            teachComment = re.search(r'点评.*',setinfor[i]).group(0)
                            flag = 3
                            find = 0
                        except AttributeError:
                            print('文件' + docFile + ' 找不到点评内容')
                
                elif flag == 3:
                    find = re.search('训练科目评定',setinfor[i])
                    if find :
                        try:
                            teacher = re.search(r'教员签字.*',setinfor[i]).group(0)
                            flag = 4
                            find = 0
                        except AttributeError:
                            print('文件' + docFile + ' 找不到教员信息')
                else:
                    find = 0
            #根据关键字查询内容
            try:
                nameNum = setinfor.index('姓名')
                name = setinfor[nameNum+1] #按行读取，所以填写内容会出现在下一项
                name = name.strip()
                if name == '':
                    continue
            except ValueError:
                print('文件' + docFile + ' 表格中未查询到填写姓名的地方')
            
            try:
                licNum = setinfor.index('执照类别')
            except ValueError:
                print('文件' + docFile + ' 表格中未查询到填写执照类别的地方')
            
            try:
                levelNum = setinfor.index('管制级别')
            except ValueError:
                print('文件' + docFile + ' 表格中未查询到填写管制级别的地方')
            
            try:
                posiNum = setinfor.index('单位及职务')
            except ValueError:
                print('文件' + docFile + ' 表格中未查询到填写单位及职务的地方')
            
            #将信息导入到列表
            #姓名
            
            diction.append(name) #姓名直接读取所在文件夹姓名，防止由于错别字导致的姓名不统一
            
            #日期格式修改：将xxxx年xx月xx日替换成xxxx/xx/xx
            time = time.replace('年','/')
            time = time.replace('月','/')
            time = time.replace('日','/')
            diction.append(time)
            
            #班组
            group_norm = os.path.normpath(docFile)
            group_norm = group_norm.split(os.sep)
            team = group_norm[-3]
            team = team.strip('班')
            team_num = chinese2digits(team)
            if team_num == -1:
                diction.append(team)
                print(group + ' ' + crew_name +'的班组无法转换成阿拉伯数字，写入默认格式的班级')
            else:
                diction.append(team_num)
                
            #见习扇区,默认为大写S
            diction.append('S')
            
            #教员
            teacher = teacher.strip('教员签字')
            teacher = teacher.strip(':')
            teacher = teacher.strip('：')
            teacher = teacher.strip()
            diction.append(teacher)
            
            #时长，默认0.75
            diction.append(0.75)
            
            #执照类别
            diction.append(setinfor[licNum+1])
            
            #管制等级
            diction.append(setinfor[levelNum+1])
            
            #单位及职务
            diction.append(setinfor[posiNum+1])
            
            #科目自我评定
            diction.append(setinfor[selfCommentNum])
            
            #教员评定
            diction.append(teachComment)
            #diction['执照类别'] = setinfor[licNum+1]
            #diction['管制级别'] = setinfor[levelNum+1]
            #diction['单位及职务'] = setinfor[posiNum+1]
            #diction['复训时间'] = time
            #diction['科目自我评定'] = setinfor[selfCommentNum]
            #diction['教员评定'] = teachComment
            
            #if ((not os.path.exists(target_file)) or (Num == 0)):
               # data_frame = pd.DataFrame(columns=seq,index=[])
                #indexsize = data_frame.index.size
                #data_frame.loc[indexsize] = diction
                #data = pd.DataFrame(diction,index=seq)
                #data_frame.to_excel(target_file, sheet_name=crew_name, index=False, header=True)
                #Num = Num + 1
                #print('成功写入 ' + group + ' ' + crew_name + '的第1行数据')
           # else: 
           
            #读取目标文件,并导出到excel
            Num = Num + 1
            write_excel_xlsx(target_file, crew_name, diction)
            print('写入 ' + group + ' ' + crew_name + '的第' + str(Num) + '条数据成功！')
        data_writer.close()
